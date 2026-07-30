import os
import io
import json
import logging
import time
import base64
import re
from datetime import datetime
from typing import List
from pydantic import BaseModel, Field
from google import genai
from google.genai import types
from tenacity import retry, stop_after_attempt, wait_exponential
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openai import OpenAI
from PIL import Image
from dotenv import load_dotenv

# Google API Imports
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload

# ==========================================================
# 🔐 1. SECURE ENVIRONMENT CONFIGURATION
# ==========================================================
load_dotenv()

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY", "").strip()
OPENROUTER_API_KEY = os.getenv("OPENROUTER_API_KEY", "").strip()

# Clean Gemini key of any hidden ASCII or white space characters
CLEAN_GEMINI_KEY = GEMINI_API_KEY.encode("ascii", "ignore").decode("ascii").strip()

# ==========================================================
# ⚙️ 2. GENERAL INVOICES PIPELINE CONFIGURATION
# ==========================================================
GENERAL_INPUT_FOLDER_ID = "12hCBCnmkRCjYdBDpQ5OFogHti8jdyGNR"
GENERAL_OUTPUT_PARENT_ID = "1p0nwqifWS85xbyJ4aFv5exfgedi5z7hI"
GENERAL_OUTPUT_FOLDER_NAME = "Automated_Invoices_Master"
GENERAL_MASTER_EXCEL_NAME = "Classified_Invoices_Report.xlsx"
GENERAL_MASTER_SHEET_ID = "1gqBbw6Do5NSHhd8uwz-9GJsIr1wcRHNu"

# ==========================================================
# ⚙️ 3. ROBOT MATERIALS PIPELINE CONFIGURATION
# ==========================================================
ROBOT_INPUT_FOLDER_ID = "1ceNx8mRmyYelUyrQswvJkCKALI7fujLh"
ROBOT_OUTPUT_PARENT_ID = "1-f3w6E5TH7Gw8Kp98jk85UHL1bSz6ueE"
ROBOT_OUTPUT_FOLDER_NAME = "Robot_Materials_Master"
ROBOT_MASTER_EXCEL_NAME = "Robot_Materials_Report.xlsx"
ROBOT_MASTER_SHEET_ID = "1xCV6zmFPlaHL3sInZLJH4HqZDd7CuKBw"

TEMP_DIR = "temp_processing"
os.makedirs(TEMP_DIR, exist_ok=True)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - [%(levelname)s] - %(message)s", datefmt="%H:%M:%S")
logger = logging.getLogger(__name__)

SCOPES = ['https://www.googleapis.com/auth/drive']

def authenticate_drive():
    """Authenticates and builds the Google Drive API service."""
    creds = None
    if os.path.exists('token.json'):
        creds = Credentials.from_authorized_user_file('token.json', SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file('credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)
        with open('token.json', 'w') as token:
            token.write(creds.to_json())
    return build('drive', 'v3', credentials=creds)

def get_or_create_output_folder(drive_service, parent_id, folder_name):
    """Finds or creates a dedicated output folder inside the specified parent folder."""
    query = f"'{parent_id}' in parents and name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = drive_service.files().list(
        q=query, fields="files(id, name)", includeItemsFromAllDrives=True, supportsAllDrives=True
    ).execute()
    folders = results.get('files', [])
    
    if folders:
        return folders[0]['id']
    
    folder_metadata = {
        'name': folder_name,
        'mimeType': 'application/vnd.google-apps.folder',
        'parents': [parent_id]
    }
    folder = drive_service.files().create(body=folder_metadata, fields='id', supportsAllDrives=True).execute()
    logger.info(f"📁 Created dedicated output folder: {folder_name}")
    return folder['id']

class InvoiceData(BaseModel):
    document_type: str = Field(description="Classify strictly as 'GST' or 'NON-GST'. CRITICAL RULE: If the bill is missing a formal GST Number AND an Invoice Number, classify it as 'NON-GST', even if the word 'GST' is written somewhere.")
    invoice_date: str = Field(description="The date of the invoice/bill strictly in YYYY-MM-DD format. If missing, return 'UNKNOWN'.")
    seller_name: str = Field(description="Name of the seller company or shop.")
    seller_gst: str = Field(description="Seller's GST Number. If not found, output 'N/A'.")
    invoice_number: str = Field(description="Invoice or Bill number. If not found, output 'N/A'.")
    products: List[str] = Field(description="A list of all product names or particulars purchased.")
    total_gst: str = Field(description="The total GST or tax amount charged (e.g., 161.18 or 4230.00). If no tax is charged, output '0'.")
    grand_total: str = Field(description="The final grand total amount billed.")

def encode_image_to_base64(image_path):
    """Compresses and converts local image to base64 for OpenRouter API to prevent payload limits."""
    with Image.open(image_path) as img:
        img = img.convert("RGB")
        img.thumbnail((800, 800))
        buffer = io.BytesIO()
        img.save(buffer, format="JPEG", quality=70)
        return base64.b64encode(buffer.getvalue()).decode('utf-8')

# 🚨 DUAL-ENGINE FALLBACK: Tries Gemini, falls back to OpenRouter automatically
@retry(stop=stop_after_attempt(5), wait=wait_exponential(multiplier=2, min=5, max=60))
def extract_data_safely(client, uploaded_file, prompt, local_file_path):
    """Extracts structured invoice data using Gemini, with automated OpenRouter fallback."""
    try:
        logger.info("Attempting extraction with Gemini API...")
        response = client.models.generate_content(
            model='gemini-3.5-flash',
            contents=[uploaded_file, prompt],
            config=types.GenerateContentConfig(
                response_mime_type="application/json", 
                response_schema=InvoiceData, 
                temperature=0.0
            ),
        )
        return response.text
        
    except Exception as e:
        logger.warning(f"⚠️ Gemini API failed/rate-limited ({e}). Switching to OpenRouter API...")
        
        openrouter_client = OpenAI(
            api_key=OPENROUTER_API_KEY,
            base_url="https://openrouter.ai/api/v1"
        )
        
        base64_image = encode_image_to_base64(local_file_path)
        
        strict_prompt = prompt + """
        Return ONLY a JSON object matching this exact schema structure:
        {
            "document_type": "GST or NON-GST",
            "invoice_date": "YYYY-MM-DD",
            "seller_name": "Name",
            "seller_gst": "GST Number or N/A",
            "invoice_number": "Number or N/A",
            "products": ["item1", "item2"],
            "total_gst": "0.00",
            "grand_total": "0.00"
        }
        """

        response = openrouter_client.chat.completions.create(
            model="nvidia/nemotron-nano-12b-v2-vl:free",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": strict_prompt},
                        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
                    ]
                }
            ],
            temperature=0.0
        )
        
        raw_text = response.choices[0].message.content.strip()
        logger.info(f"OpenRouter Raw Response: {raw_text}")
        
        # 🔍 Extraction Guardrail: Parse specifically between curly braces
        json_match = re.search(r'\{.*\}', raw_text, re.DOTALL)
        if json_match:
            return json_match.group(0)
            
        return raw_text

def ensure_clean_sheet(ws):
    """Formats header row for a sheet tab with standard styling."""
    headers = ["Date", "Seller Name", "GST Number", "Invoice Number", "Products", "Total GST", "Grand Total", "Source File"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

def process_document(file_path, master_excel_path, is_robot_pipeline=False):
    """Processes image file through Vision AI and appends extracted data into Excel report."""
    raw_filename = os.path.basename(file_path)

    client = genai.Client(api_key=CLEAN_GEMINI_KEY)
    logger.info(f"Uploading {raw_filename} to Gemini...")
    
    try:
        uploaded_file = client.files.upload(file=file_path)
    except Exception as e:
        logger.error(f"Gemini upload error: {e}")
        uploaded_file = "Direct OpenRouter Fallback"

    prompt = """
    You are an expert accountant AI. 
    Analyze the attached bill/invoice and extract exactly the requested details.
    
    CRITICAL DATE RULE: Pay VERY close attention to handwritten dates. Assume dates are written in standard Indian format (DD/MM/YY or DD/MM/YYYY). For example, a handwritten "16/6/26" must be interpreted as June 16, 2026, NOT November.
    
    CRITICAL GST RULE: 
    1. Look closely at the SGST and CGST values. Be careful with separate 'Rs' and 'P' (Paise) columns (e.g., a 10 under Rs and an 80 under P means 10.80). 
    2. The 'total_gst' MUST be the exact mathematical sum of all taxes (SGST + CGST + IGST). Do NOT confuse the tax percentage (e.g., 9%) with the actual tax amount.
    3. To guarantee accuracy, mathematically cross-check it: (Grand Total - Subtotal = Total GST).
    
    CRITICAL GSTIN RULE: Distinguish between the Seller (Supplier) GSTIN and the Buyer/Consignee GSTIN. Look for labels like 'GSTIN' or 'GSTIN/UIN' located near the top header next to the seller's company name. Do not output the buyer's GSTIN; always extract the seller/supplier's GSTIN. If truly not found, output 'N/A'.
    
    ONLY extract the date, seller name, GST number, invoice number, product names, total GST, and grand total.
    """

    try:
        response_text = extract_data_safely(client, uploaded_file, prompt, file_path)
        structured_data = InvoiceData(**json.loads(response_text))
        
        try:
            dt = datetime.strptime(structured_data.invoice_date, "%Y-%m-%d")
            month_name = dt.strftime("%b")
        except:
            month_name = "Unknown"

        # Universal GST vs Non-GST Classification
        gst_is_missing = structured_data.seller_gst.strip().upper() in ["N/A", "UNKNOWN", "NONE", ""]
        inv_is_missing = structured_data.invoice_number.strip().upper() in ["N/A", "UNKNOWN", "NONE", ""]
        base_type = "Non-GST" if (gst_is_missing and inv_is_missing) or "NON" in structured_data.document_type.upper() else "GST"
        
        # Route to specific sheet tab name
        if is_robot_pipeline:
            target_sheet_name = f"Robot {base_type} {month_name}"
        else:
            target_sheet_name = f"{base_type} {month_name}"

        if os.path.exists(master_excel_path):
            wb = load_workbook(master_excel_path)
        else:
            wb = Workbook()

        if target_sheet_name not in wb.sheetnames:
            wb.create_sheet(target_sheet_name)
            
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1 and target_sheet_name != "Sheet": 
            del wb["Sheet"]

        ws = wb[target_sheet_name]

        ensure_clean_sheet(ws)

        products_string = ", ".join(structured_data.products)

        new_row = [
            structured_data.invoice_date,
            structured_data.seller_name,
            structured_data.seller_gst,
            structured_data.invoice_number,
            products_string,
            structured_data.total_gst,
            structured_data.grand_total,
            raw_filename
        ]

        # Read existing data rows
        valid_rows = []
        if ws.max_row > 1:
            for r in range(2, ws.max_row + 1):
                row_vals = [ws.cell(row=r, column=c).value for c in range(1, 9)]
                if row_vals[0] is not None and row_vals[0] != "Date":
                    valid_rows.append(row_vals)

        # Append new row and sort chronologically by date
        valid_rows.append(new_row)
        valid_rows.sort(key=lambda x: str(x[0]))

        # Rewrite sheet vertically starting from Row 2
        ws.delete_rows(2, ws.max_row)

        na_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row_idx, r_vals in enumerate(valid_rows, start=2):
            for col_idx, val in enumerate(r_vals, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                if str(val).strip().upper() == "N/A":
                    cell.fill = na_fill

        # Auto-adjust column widths dynamically
        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 60)
            
        wb.save(master_excel_path)
        logger.info(f"✅ Appended vertically and saved to sheet: {target_sheet_name}.")
        return True

    except Exception as e:
        logger.error(f"Extraction Error: {str(e)}")
        return False
    finally:
        try:
            if uploaded_file != "Direct OpenRouter Fallback":
                client.files.delete(name=uploaded_file.name)
        except: pass

def check_and_process_folder(drive_service, input_folder_id, parent_output_id, output_folder_name, master_excel_name, is_robot=False):
    """Scans Google Drive input folder, downloads new files, runs extraction, and syncs Excel output back to Drive."""
    query = f"'{input_folder_id}' in parents and trashed=false"
    results = drive_service.files().list(
        q=query, fields="files(id, name)", includeItemsFromAllDrives=True, supportsAllDrives=True
    ).execute()
    
    items = results.get('files', [])

    for item in items:
        file_id = item['id']
        file_name = item['name']

        if file_name.endswith('.completed') or file_name.endswith('.error'):
            continue
        if not file_name.lower().endswith(('.pdf', '.jpg', '.jpeg', '.png')):
            continue

        prefix = "🤖 [ROBOT]" if is_robot else "📄 [GENERAL]"
        logger.info(f"--- {prefix} Found new file: {file_name} ---")
        
        local_path = os.path.join(TEMP_DIR, file_name)
        request = drive_service.files().get_media(fileId=file_id, supportsAllDrives=True)
        fh = io.FileIO(local_path, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()

        output_folder_id = get_or_create_output_folder(drive_service, parent_output_id, output_folder_name)
        master_local_path = os.path.join(TEMP_DIR, master_excel_name)
        
        master_query = f"'{output_folder_id}' in parents and name='{master_excel_name}' and trashed=false"
        master_results = drive_service.files().list(
            q=master_query, fields="files(id, name)", includeItemsFromAllDrives=True, supportsAllDrives=True
        ).execute()
        
        master_items = master_results.get('files', [])
        master_file_id = None
        
        if master_items:
            master_file_id = master_items[0]['id']
            master_req = drive_service.files().get_media(fileId=master_file_id, supportsAllDrives=True)
            master_fh = io.FileIO(master_local_path, 'wb')
            master_dl = MediaIoBaseDownload(master_fh, master_req)
            master_done = False
            while not master_done:
                _, master_done = master_dl.next_chunk()
        else:
            if os.path.exists(master_local_path):
                os.remove(master_local_path)

        success = process_document(local_path, master_local_path, is_robot_pipeline=is_robot)

        if success:
            media = MediaFileUpload(master_local_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            
            if master_file_id:
                drive_service.files().update(
                    fileId=master_file_id, media_body=media, supportsAllDrives=True
                ).execute()
            else:
                file_metadata = {'name': master_excel_name, 'parents': [output_folder_id]}
                drive_service.files().create(
                    body=file_metadata, media_body=media, fields='id', supportsAllDrives=True
                ).execute()

            drive_service.files().update(fileId=file_id, body={'name': file_name + '.completed'}, supportsAllDrives=True).execute()
            logger.info(f"✅ Success: {file_name} marked as .completed in Drive.")
        else:
            drive_service.files().update(fileId=file_id, body={'name': file_name + '.error'}, supportsAllDrives=True).execute()
            logger.error(f"❌ Failed: {file_name} marked as .error in Drive.")
        
        try: os.remove(local_path)
        except: pass
        try: os.remove(master_local_path)
        except: pass

def autonomous_cloud_watcher():
    """Main continuous watchdog loop monitoring both General and Robot input pipelines."""
    logger.info("🔐 Authenticating with Google Drive API...")
    drive_service = authenticate_drive()
    logger.info("🚀 Cloud Engine Initialized. Monitoring BOTH General and Robot folders...")
    
    while True:
        try:
            # 1. Process General Invoices Pipeline
            check_and_process_folder(
                drive_service, 
                GENERAL_INPUT_FOLDER_ID, 
                GENERAL_OUTPUT_PARENT_ID, 
                GENERAL_OUTPUT_FOLDER_NAME, 
                GENERAL_MASTER_EXCEL_NAME, 
                is_robot=False
            )

            # 2. Process Robot Materials Pipeline
            check_and_process_folder(
                drive_service, 
                ROBOT_INPUT_FOLDER_ID, 
                ROBOT_OUTPUT_PARENT_ID, 
                ROBOT_OUTPUT_FOLDER_NAME, 
                ROBOT_MASTER_EXCEL_NAME, 
                is_robot=True
            )

        except Exception as e:
            logger.error(f"Watchdog error: {str(e)}")
            
        time.sleep(10)

if __name__ == "__main__":
    autonomous_cloud_watcher()