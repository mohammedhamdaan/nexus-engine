import os
import sys
import json
import shutil
import platform
import subprocess
import traceback
import openpyxl
from fastapi import FastAPI, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from agent import (
    process_document, 
    authenticate_drive, 
    TEMP_DIR,
    GENERAL_MASTER_EXCEL_NAME,
    ROBOT_MASTER_EXCEL_NAME,
    GENERAL_OUTPUT_PARENT_ID,
    ROBOT_OUTPUT_PARENT_ID
)

app = FastAPI()

PIPELINES_FILE = "pipelines.json"

# --- THE BULLETPROOF BROWSER FIX ---
def force_open_browser(url):
    """Safely forces the OS to open a link even when running in a --windowed PyInstaller app."""
    try:
        if platform.system() == 'Windows':
            os.startfile(url) # The native Windows Shell command
        elif platform.system() == 'Darwin':
            subprocess.Popen(['open', url]) # The native Mac command
        else:
            subprocess.Popen(['xdg-open', url]) # Linux fallback
    except Exception as e:
        print(f"Failed to open browser: {e}")
# -----------------------------------

def load_pipelines():
    if os.path.exists(PIPELINES_FILE):
        with open(PIPELINES_FILE, "r") as f:
            return json.load(f)
    else:
        default_pipelines = {
            "general": {
                "id": "general",
                "name": "General",
                "folder_id": GENERAL_OUTPUT_PARENT_ID,
                "master_name": GENERAL_MASTER_EXCEL_NAME,
                "is_robot": False
            },
            "robot": {
                "id": "robot",
                "name": "Robot",
                "folder_id": ROBOT_OUTPUT_PARENT_ID,
                "master_name": ROBOT_MASTER_EXCEL_NAME,
                "is_robot": True
            }
        }
        with open(PIPELINES_FILE, "w") as f:
            json.dump(default_pipelines, f, indent=4)
        return default_pipelines

def save_pipelines(pipelines):
    with open(PIPELINES_FILE, "w") as f:
        json.dump(pipelines, f, indent=4)

class PipelineRequest(BaseModel):
    name: str

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://127.0.0.1:8000"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/pipelines")
async def get_pipelines():
    return load_pipelines()

@app.post("/create_pipeline")
async def create_pipeline(req: PipelineRequest):
    pipelines = load_pipelines()
    pipeline_id = req.name.lower().replace(" ", "_")
    
    if pipeline_id in pipelines:
        return {"status": "error", "message": "Folder already exists!"}

    try:
        drive_service = authenticate_drive()
        
        # The specific "AI Agents" folder ID you provided
        AI_AGENTS_FOLDER_ID = "1PijHV-R2E7KiAVI-fKPCN0cX2e9113XJ"
        
        # 1. Create the main new folder (e.g., "Ammar") INSIDE "AI Agents"
        main_folder_metadata = {
            'name': req.name,
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [AI_AGENTS_FOLDER_ID]
        }
        main_folder = drive_service.files().create(body=main_folder_metadata, fields='id', supportsAllDrives=True).execute()
        main_folder_id = main_folder.get('id')

        # 2. Create the "input" subfolder
        input_folder_metadata = {
            'name': 'input',
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [main_folder_id]
        }
        input_folder = drive_service.files().create(body=input_folder_metadata, fields='id', supportsAllDrives=True).execute()
        input_folder_id = input_folder.get('id')

        # 3. Create the "output" subfolder
        output_folder_metadata = {
            'name': 'output',
            'mimeType': 'application/vnd.google-apps.folder',
            'parents': [main_folder_id]
        }
        output_folder = drive_service.files().create(body=output_folder_metadata, fields='id', supportsAllDrives=True).execute()
        output_folder_id = output_folder.get('id')

        # Add the new pipeline to our tracker
        pipelines[pipeline_id] = {
            "id": pipeline_id,
            "name": req.name,
            "folder_id": output_folder_id,
            "input_folder_id": input_folder_id,
            "master_name": f"{req.name}_Report.xlsx", 
            "is_robot": False  
        }
        save_pipelines(pipelines)
        return {"status": "success", "pipeline": pipelines[pipeline_id]}
        
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/upload")
async def upload_file(file: UploadFile = File(...), pipeline: str = Form(...)):
    pipelines = load_pipelines()
    if pipeline not in pipelines:
        return {"status": "error", "message": f"Invalid folder: {pipeline}"}
        
    p_config = pipelines[pipeline]
    is_robot = p_config.get("is_robot", False)
    master_name = p_config["master_name"]
    output_parent_id = p_config["folder_id"]
    
    # 1. Define Paths
    save_path = os.path.join(TEMP_DIR, file.filename)
    master_path = os.path.join("output_excel", master_name)
    os.makedirs(TEMP_DIR, exist_ok=True)
    os.makedirs("output_excel", exist_ok=True)
    
    # 2. Save incoming image
    with open(save_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    try:
        drive_service = authenticate_drive()
        
        # Upload the original invoice to the "input" folder if it exists
        input_folder_id = p_config.get("input_folder_id")
        if input_folder_id:
            try:
                img_media = MediaFileUpload(save_path, mimetype=file.content_type)
                drive_service.files().create(
                    body={'name': file.filename, 'parents': [input_folder_id]},
                    media_body=img_media,
                    supportsAllDrives=True
                ).execute()
            except Exception as img_err:
                print(f"⚠️ Failed to upload image to input folder: {img_err}")
                
        file_id = None
        sheet_url = None
        
        # 3. Resolve the Google Sheet ID dynamically or use hardcoded IDs
        if pipeline == "general":
            file_id = "1gqBbw6Do5NSHhd8uwz-9GJsIr1wcRHNu" 
        elif pipeline == "robot":
            file_id = "1xCV6zmFPlaHL3sInZLJH4HqZDd7CuKBw"
        else:
            # Custom folder: Search inside the Drive folder automatically
            query = f"name='{master_name}' and '{output_parent_id}' in parents and trashed=false"
            results = drive_service.files().list(
                q=query, 
                fields="files(id, webViewLink)", 
                supportsAllDrives=True
            ).execute()
            items = results.get('files', [])
            
            if items:
                file_id = items[0]['id']
                sheet_url = items[0].get('webViewLink')

        # 4. Download if the file exists, otherwise create the perfect template!
        if file_id:
            request = drive_service.files().get_media(fileId=file_id, supportsAllDrives=True)
            with open(master_path, 'wb') as f:
                downloader = MediaIoBaseDownload(f, request)
                done = False
                while not done:
                    _, done = downloader.next_chunk()
            
            if not sheet_url:
                file_info = drive_service.files().get(fileId=file_id, fields="webViewLink", supportsAllDrives=True).execute()
                sheet_url = file_info.get("webViewLink")
        else:
            if os.path.exists(master_path): os.remove(master_path)
            
            # --- THE FINAL FIX FOR BRAND NEW CUSTOM FOLDERS ---
            # AI Agents crash if they don't see expected columns/headers in the Excel sheet.
            # We fix this by downloading the working "General" sheet, wiping its old data,
            # and feeding it to the AI as a perfect blank template with all the correct headers!
            try:
                general_id = "1gqBbw6Do5NSHhd8uwz-9GJsIr1wcRHNu"
                request = drive_service.files().get_media(fileId=general_id, supportsAllDrives=True)
                with open(master_path, 'wb') as f:
                    downloader = MediaIoBaseDownload(f, request)
                    done = False
                    while not done:
                        _, done = downloader.next_chunk()
                
                # Now we wipe the old data, leaving only the headers in Row 1
                wb = openpyxl.load_workbook(master_path)
                ws = wb.active
                if ws.max_row > 1:
                    ws.delete_rows(2, ws.max_row - 1)
                wb.save(master_path)
            except Exception as template_err:
                print(f"Failed to clone template: {template_err}")
                wb = openpyxl.Workbook()
                wb.save(master_path)
            # -------------------------------------------------

        # 5. Process the document using the Agent
        success = process_document(save_path, master_path, is_robot_pipeline=is_robot)
        
        # 6. Upload back to Drive
        if success and os.path.exists(master_path):
            media = MediaFileUpload(master_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            
            if file_id:
                # Update existing
                drive_service.files().update(
                    fileId=file_id, 
                    media_body=media, 
                    supportsAllDrives=True
                ).execute()
            else:
                # Create brand new sheet
                file_metadata = {'name': master_name, 'parents': [output_parent_id]}
                new_file = drive_service.files().create(
                    body=file_metadata, 
                    media_body=media, 
                    supportsAllDrives=True, 
                    fields="id, webViewLink"
                ).execute()
                sheet_url = new_file.get("webViewLink")
            
            # Cleanup
            if os.path.exists(save_path): os.remove(save_path)
            
            # Use the bulletproof Windows native command to open the sheet immediately
            if sheet_url:
                force_open_browser(sheet_url)
            
            return {"status": "success", "pipeline": pipeline, "sheet_url": sheet_url}
        else:
            if os.path.exists(save_path): os.remove(save_path)
            return {"status": "error", "message": "Extraction or local save failed."}

    except Exception as e:
        # If it crashes, immediately write a log file so we know exactly why
        with open("crash_log.txt", "w") as f:
            f.write(traceback.format_exc())
        if os.path.exists(save_path): os.remove(save_path)
        return {"status": "error", "message": f"Process failed: {str(e)}"}